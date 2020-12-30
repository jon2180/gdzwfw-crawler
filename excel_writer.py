"""
作用：用于把具体数据写入到 excel 中

写入逻辑：读取模板 excel 前两行的数据，复制粘贴到欲生成的 excel 中，之后每写入一行，就会使行标加1，让下一次写入的时候能够写在下一行，
最后使用 save 方法，并参入生成的路径，输出文件(这样是为了不破坏模板文件)
"""
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet import worksheet

from model import PowerAndResponsibility


class PARExcelWriter:
    workbook: Workbook
    current_row_index: 'int > 0' = 1
    template_xlsx: str = None

    def __init__(
            self,
            template_path: str,
    ):
        """
        :keyword:

        :param template_path: 文件路径
        :param mode: 文件写入模式: 'w' 直接写在模板文件中写， 'cw' copy and write copy 模板文件的前两行，然后在第三行开始写入
        """
        self.template_xlsx = template_path
        self.current_row_index = 1
        self.workbook = Workbook()
        self._init_workbook()

    def _init_workbook(self):
        """
        把模板的前两行给复制到产出 excel 文件中
        :return:
        """
        curr_wb = load_workbook(self.template_xlsx)
        curr_ws = curr_wb.active

        ws = self.workbook.active
        for row_index in range(1, 3):
            for col_index in range(1, 57):
                ws.cell(row_index, col_index, curr_ws.cell(row_index, col_index).value)
        self.current_row_index += 2
        curr_wb.close()

    def write_row(self, data: PowerAndResponsibility):
        work_sheet = self.workbook.active

        assert isinstance(work_sheet, worksheet.Worksheet)
        assert isinstance(self.current_row_index, int)
        assert isinstance(data, PowerAndResponsibility)

        # 标题
        work_sheet[f'A{self.current_row_index}'] = data.libNum or ''
        work_sheet[f'B{self.current_row_index}'] = data.catalogId or ''
        work_sheet[f'C{self.current_row_index}'] = data.title or ''
        work_sheet[f'D{self.current_row_index}'] = data.issuedNum or ''
        work_sheet[f'E{self.current_row_index}'] = data.publisher or '' # 发布者

        # 发布日期
        work_sheet[f'F{self.current_row_index}'] = data.publishDate or datetime.now().strftime('%d/%m/%Y')
        work_sheet[f'G{self.current_row_index}'] = data.source or '广东省政务服务网'
        work_sheet[f'H{self.current_row_index}'] = data.belongDomain or ''
        work_sheet[f'I{self.current_row_index}'] = data.genre or '清单'
        work_sheet[f'J{self.current_row_index}'] = data.classify or ''

        # k 摘要，省略
        work_sheet[f'L{self.current_row_index}'] = data.originalAddress or ''
        work_sheet[f'M{self.current_row_index}'] = data.htmlContent or ''
        work_sheet[f'N{self.current_row_index}'] = data.creatorName or ''
        work_sheet[f'O{self.current_row_index}'] = data.creatorName or ''

        if data.pk_field is not None:
            work_sheet[f'P{self.current_row_index}'] = data.pk_field
        if data.pk_industry is not None:
            work_sheet[f'Q{self.current_row_index}'] = data.pk_industry
        if data.pk_fast_time is not None:
            work_sheet[f'R{self.current_row_index}'] = data.pk_fast_time
        if data.pk_region is not None:
            work_sheet[f'S{self.current_row_index}'] = data.pk_region
        if data.pk_item_type is not None:
            work_sheet[f'T{self.current_row_index}'] = data.pk_item_type

        if data.pk_basis is not None:
            work_sheet[f'U{self.current_row_index}'] = data.pk_basis
        if data.pk_exe_level is not None:
            work_sheet[f'V{self.current_row_index}'] = data.pk_exe_level
        if data.pk_item_source is not None:
            work_sheet[f'W{self.current_row_index}'] = data.pk_item_source
        if data.pk_org_name is not None:
            work_sheet[f'X{self.current_row_index}'] = data.pk_org_name
        if data.pk_mplementer_kind is not None:
            work_sheet[f'Y{self.current_row_index}'] = data.pk_mplementer_kind

        if data.pk_declare_type is not None:
            work_sheet[f'Z{self.current_row_index}'] = data.pk_declare_type

        if data.pk_commitment_time is not None:
            work_sheet[f'AA{self.current_row_index}'] = data.pk_commitment_time
        if data.pk_legal_time is not None:
            work_sheet[f'AB{self.current_row_index}'] = data.pk_legal_time
        if data.pk_presence_num is not None:
            work_sheet[f'AC{self.current_row_index}'] = data.pk_presence_num
        if data.pk_presence_reason is not None:
            work_sheet[f'AD{self.current_row_index}'] = data.pk_presence_reason
        if data.pk_service_objects is not None:
            work_sheet[f'AE{self.current_row_index}'] = data.pk_service_objects

        if data.pk_operation_scope is not None:
            work_sheet[f'AF{self.current_row_index}'] = data.pk_operation_scope
        if data.pk_exercise_content is not None:
            work_sheet[f'AG{self.current_row_index}'] = data.pk_exercise_content
        if data.pk_implementation_object is not None:
            work_sheet[f'AH{self.current_row_index}'] = data.pk_implementation_object
        if data.pk_requirements is not None:
            work_sheet[f'AI{self.current_row_index}'] = data.pk_requirements
        if data.pk_approval_mode is not None:
            work_sheet[f'AJ{self.current_row_index}'] = data.pk_approval_mode

        if data.pk_materials is not None:
            work_sheet[f'AK{self.current_row_index}'] = data.pk_materials
        if data.pk_process_form is not None:
            work_sheet[f'AL{self.current_row_index}'] = data.pk_process_form
        if data.pk_procedures is not None:
            work_sheet[f'AM{self.current_row_index}'] = data.pk_procedures
        if data.pk_process_time is not None:
            work_sheet[f'AN{self.current_row_index}'] = data.pk_process_time
        if data.pk_process_place is not None:
            work_sheet[f'AO{self.current_row_index}'] = data.pk_process_place

        if data.pk_result_name is not None:
            work_sheet[f'AP{self.current_row_index}'] = data.pk_result_name
        if data.pk_result_sample is not None:
            work_sheet[f'AQ{self.current_row_index}'] = data.pk_result_sample
        if data.pk_result_type is not None:
            work_sheet[f'AR{self.current_row_index}'] = data.pk_result_type
        if data.pk_result_collection_method is not None:
            work_sheet[f'AS{self.current_row_index}'] = data.pk_result_collection_method
        if data.pk_logistics_express is not None:
            work_sheet[f'AT{self.current_row_index}'] = data.pk_logistics_express

        if data.pk_online_payment is not None:
            work_sheet[f'AU{self.current_row_index}'] = data.pk_online_payment
        if data.pk_online_processing is not None:
            work_sheet[f'AV{self.current_row_index}'] = data.pk_online_processing
        if data.pk_online_processing_depth is not None:
            work_sheet[f'AW{self.current_row_index}'] = data.pk_online_processing_depth
        if data.pk_appointment is not None:
            work_sheet[f'AX{self.current_row_index}'] = data.pk_appointment
        if data.pk_is_charge is not None:
            work_sheet[f'AY{self.current_row_index}'] = data.pk_is_charge

        if data.pk_charge_standard is not None:
            work_sheet[f'AZ{self.current_row_index}'] = data.pk_charge_standard

        if data.pk_intermediary is not None:
            work_sheet[f'BA{self.current_row_index}'] = data.pk_intermediary
        if data.pk_special_procedure is not None:
            work_sheet[f'BB{self.current_row_index}'] = data.pk_special_procedure
        if data.pk_consultation_ways is not None:
            work_sheet[f'BC{self.current_row_index}'] = data.pk_consultation_ways
        if data.pk_monitor_ways is not None:
            work_sheet[f'BD{self.current_row_index}'] = data.pk_monitor_ways

        self.current_row_index += 1

    def save(self, out_xlsx: str):
        self.workbook.save(out_xlsx)
